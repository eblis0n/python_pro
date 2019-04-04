# encoding=utf-8

import unittest
from utils.methods import set_dp_interface as set_dp_interface
import constants.deploy as deploy
import pymysql
import os
import configparser  as cparser
import openpyxl

############################### 配置文件的地址######################

# base_dr = str(os.path.dirname(os.path.dirname(__file__)))
# bae_idr = base_dr.replace('\\', '/')
#
# file_path = bae_idr + "/config.ini"
#
# conf = cparser.ConfigParser()
#
# conf.read(file_path)
#
#
# # ######################读取配置文件上###################################
# mysqlHost = conf.get("mysqlconf", "host")   # 读取配置文件上数据库主机号
# mysqlPort = conf.get("mysqlconf", "port")    # 读取配置文件数据库的端口号
# mysqlUser = conf.get("mysqlconf", "user")    # 读取配置文件数据库的账号
# mysqlPassword = conf.get("mysqlconf", "password")   # 读取配置文件数据库的密码
# mysqlName = conf.get("mysqlconf", "db_name")   # 读取配置文件数据库的名字
# wbook = openpyxl.load_workbook(conf.get("excel", "excel_path")) # 读取配置文件的Excel文档地址
# table2 = wbook[conf.get("excel", "table")] # 读取配置文件的Excel文档所使用的用例table
#######################配置Excel的用例地址#################################
# 直接读取不利于后期修改Excel文档存放变更
# wbook=openpyxl.load_workbook(bae_idr + "/testexcel/ssptest.xlsx")
# table2=wbook['Test_Case2']


#######################从deploy里面读取配置文件内容#################################

mysqlHost = deploy.mysqlHost
mysqlPort = deploy.mysqlPort
mysqlUser = deploy.mysqlUser
mysqlPassword = deploy.mysqlPassword
mysqlName = deploy.mysqlName
wbook = deploy.wbook
table2 = deploy.table2

#######################根据Excel读取本接口对应的key#################################
begtest=[]
endtest=[]

for i in range(1,int(table2.max_column)):
    if table2["C"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

for j in range(1, len(begtest)):
    for k in range(2, int(begtest[j] + 1)):
        # print("excel:" + table2["C" + str(k)].value)
        if os.path.basename(__file__)[0:-3] == table2["D" + str(k)].value:

            # 接口参数
            queryentry_url = table2["E" + str(k)].value # 接口路径
            sspFinanceAnalysisPlaySummary = table2["F" + str(k)].value # 接口地址
            analysisPlaySummary_data = table2["G" + str(k)].value # 请求参数

            # 数据库条件
            media_provider_id = table2["H" + str(k)].value
            # create_time = table2["I" + str(k)].value
            # transaction_date = table2["H" + str(k)].value
            startDate = table2["I" + str(k)].value
            endDate = table2["J" + str(k)].value
            limits = table2["k" + str(k)].value
            break
########################################################

# 对应的接口：/ssp/finance/analysis/play/summary
class test_analysis_play_summary (unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pymysqlconn = pymysql.connect(host=mysqlHost, port=int(mysqlPort), user=mysqlUser, passwd=mysqlPassword, db=mysqlName,
                           charset="utf8",cursorclass = pymysql.cursors.DictCursor)
        cls.pymysqlcursor = cls.pymysqlconn.cursor()

        # 接口
        # cls.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlaySummary, data=analysisPlaySummary_data)
        print("test_analysis_play_summary接口测试开始\n")

    @classmethod
    def tearDownClass(cls):
        print("测试结束")
        cls.pymysqlconn.close()

    # 数据汇总_请求屏数
    def test_requestScreenCount(self):
        print("开始测试test_screenCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlaySummary, data=analysisPlaySummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_requestScreenCount =  self.res_date['data'].get("requestScreenCount")
        # 执行SQL查询
        # self.sql_screenCount = get_media_screenCount(2028, '2019-02-14',conn = self.pymysqlconn )
        sql = "SELECT sum(DISTINCT request_screens) AS request_screens FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}' GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        request_screens=self.pymysqlcursor.fetchall()
        sql_request_screens = request_screens[0].get("request_screens")
        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(str(inf_requestScreenCount),str(sql_request_screens)))
        self.assertEquals(str(inf_requestScreenCount), str(sql_request_screens),"test_requestScreenCount,数据对不上" )

    # 数据汇总_展现广告屏数
    def test_validScreenCount(self):
        print("开始测试test_validscreencount")
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlaySummary, data=analysisPlaySummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_validscreencount =  self.res_date['data'].get("validScreenCount")
        # 执行SQL查询
        sql = "SELECT  sum(DISTINCT play_screens) AS play_screens FROM ssp_play_summary WHERE ssp_id = '{0}' AND transaction_date BETWEEN '{1}' AND '{2}' GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        play_screens=self.pymysqlcursor.fetchall()
        slq_play_screens =  play_screens[0].get("play_screens")
        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(str(inf_validscreencount), str(slq_play_screens)))
        self.assertEquals(str(inf_validscreencount), str(slq_play_screens), "test_validScreenCount,数据对不上" )

    # 数据汇总_请求广告次数
    def test_requestCount(self):
        print("开始测试test_requestCount")
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlaySummary, data=analysisPlaySummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_requestCount =  self.res_date['data'].get("requestCount")

        # 执行SQL查询
        sql = "SELECT sum(request_times) AS request_times FROM ssp_play_summary WHERE ssp_id = '{0}' AND transaction_date BETWEEN '{1}' AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        request_times=self.pymysqlcursor.fetchall()
        slq_request_times =  request_times[0].get("request_times")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(str(inf_requestCount), str(slq_request_times)))
        self.assertEquals(str(inf_requestCount), str(slq_request_times), "数据对不上" )

    # 数据汇总_有效播放次数
    def test_materialPlayCount(self):
        print("开始测试test_materialPlayCount")
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlaySummary, data=analysisPlaySummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialPlayCount =  self.res_date['data'].get("materialPlayCount")

        # 执行SQL查询
        sql = "SELECT sum(play_times) AS play_times FROM ssp_play_summary WHERE ssp_id = '{0}' AND transaction_date BETWEEN '{1}'AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        play_times=self.pymysqlcursor.fetchall()
        slq_play_times =  play_times[0].get("play_times")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(str(inf_materialPlayCount), str(slq_play_times)))
        self.assertEquals(str(inf_materialPlayCount), str(slq_play_times), "test_materialPlayCount数据对不上" )





if __name__ == "__main__":
    unittest.main()