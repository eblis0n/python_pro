# encoding=utf-8

import unittest
from utils.methods import skip_dependon as skip_dependon
from utils.methods import set_dp_interface as set_dp_interface
from utils.methods import comparison_1 as comparison_1
import constants.deploy as deploy
import pymysql
import os
import configparser as cparser
import openpyxl

############################### 配置文件的地址######################
#
# base_dr = str(os.path.dirname(os.path.dirname(__file__)))
# bae_idr = base_dr.replace('\\', '/')
#
# file_path = bae_idr + "/config.ini"
#
# conf = cparser.ConfigParser()
#
# conf.read(file_path)
#
#
# # ######################读取配置文件上###################################
# mysqlHost = conf.get("mysqlconf", "host")   # 读取配置文件上数据库主机号
# mysqlPort = conf.get("mysqlconf", "port")    # 读取配置文件数据库的端口号
# mysqlUser = conf.get("mysqlconf", "user")    # 读取配置文件数据库的账号
# mysqlPassword = conf.get("mysqlconf", "password")   # 读取配置文件数据库的密码
# mysqlName = conf.get("mysqlconf", "db_name")   # 读取配置文件数据库的名字
# wbook = openpyxl.load_workbook(conf.get("excel", "excel_path")) # 读取配置文件的Excel文档地址
# table2 = wbook[conf.get("excel", "table")] # 读取配置文件的Excel文档所使用的用例table
#######################配置Excel的用例地址#################################
# 直接读取不利于后期修改Excel文档存放变更
# wbook=openpyxl.load_workbook(bae_idr + "/testexcel/ssptest.xlsx")
# table2=wbook['Test_Case2']

#######################从deploy里面读取配置文件内容#################################

mysqlHost = deploy.mysqlHost
mysqlPort = deploy.mysqlPort
mysqlUser = deploy.mysqlUser
mysqlPassword = deploy.mysqlPassword
mysqlName = deploy.mysqlName
wbook = deploy.wbook
table2 = deploy.table2

#######################根据Excel读取本接口对应的key#################################
begtest=[]
endtest=[]

for i in range(1,int(table2.max_column)):
    if table2["C"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

for j in range(1, len(begtest)):
    for k in range(2, int(begtest[j] + 1)):
        # print("excel:" + table2["C" + str(k)].value)
        if os.path.basename(__file__)[0:-3] == table2["D" + str(k)].value:

            # 接口参数
            queryentry_url = table2["E" + str(k)].value # 接口路径
            sspFinanceAnalysisPlayList = table2["F" + str(k)].value # 接口地址
            FinanceAnalysisPlayList_data = table2["G" + str(k)].value # 请求参数

            # 数据库条件
            media_provider_id = table2["H" + str(k)].value
            # create_time = table2["I" + str(k)].value
            # transaction_date = table2["H" + str(k)].value
            startDate = table2["I" + str(k)].value
            endDate = table2["J" + str(k)].value
            limits = table2["k" + str(k)].value
            break
########################################################

# 对应的接口：/ssp/finance/analysis/play/list
class test_queryentry_ssp_analysisPlayList_Diagram (unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pymysqlconn = pymysql.connect(host=mysqlHost, port=int(mysqlPort), user=mysqlUser, passwd=mysqlPassword, db=mysqlName,
                           charset="utf8",cursorclass = pymysql.cursors.DictCursor)
        cls.pymysqlcursor = cls.pymysqlconn.cursor()
        # cls.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        print("queryentry_ssp_analysisPlayList_Diagram(播放趋势图)接口测试开始\n")

    @classmethod
    def tearDownClass(cls):
        print("测试结束")
        cls.pymysqlconn.close()

    # 播放趋势图统计
    def test_analysisPlayLen(self):
        print("开始测试test_analysisPlayLen")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        dataLen =  self.res_date['data']
        # 执行SQL查询
        sql = "SELECT * FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}' GROUP BY transaction_date;".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_dataLen=self.pymysqlcursor.fetchall()
        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(len(dataLen),len(sql_dataLen)))
        self.assertEqual(len(dataLen), len(sql_dataLen),"test_dataLen,数据对不上" )

    # 播放趋势图_展现广告屏数
    @skip_dependon(depend="test_analysisPlayLen")
    def test_playListValidScreenCount(self):
        print("开始测试test_playListValidScreenCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_validScreenCount = self.res_date['data']

        # 执行SQL查询
        sql = "SELECT transaction_date,SUM(play_screens) AS play_screens FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}'AND '{2}' GROUP BY transaction_date;".format( media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_play_screens = self.pymysqlcursor.fetchall()

        print("展现广告屏数比对")
        # for i in range(len(validScreenCount)):
        #     statisGroup = validScreenCount[i].get("statisGroup")
        #     inf_validScreenCount = validScreenCount[i].get("validScreenCount")
        #     sql_play_screens = play_screens[i].get("play_screens")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisGroup, inf_validScreenCount,sql_play_screens))
        # self.assertEqual(inf_validScreenCount,sql_play_screens,"test_validScreenCount,数据对不上" )
        statisDates = 'statisGroup'
        com_validScreenCount = 'validScreenCount'
        com_play_screens = 'play_screens'
        comparison_1(self, inf_validScreenCount, sql_play_screens, statisDates, com_validScreenCount, com_play_screens)

    # 播放趋势图_有效播放次数
    @skip_dependon(depend="test_analysisPlayLen")
    def test_playListplayTimes(self):
        print("开始测试test_playListplayTimes")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialPlayCount = self.res_date['data']

        # 执行SQL查询
        sql = "SELECT transaction_date,SUM(play_times) AS play_times FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}'AND '{2}' GROUP BY transaction_date;".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_play_times = self.pymysqlcursor.fetchall()

        print("有效播放次数比对")
        # for i in range(len(materialPlayCount)):
        #     statisGroup = materialPlayCount[i].get("statisGroup")
        #     inf_materialPlayCount = materialPlayCount[i].get("materialPlayCount")
        #     sql_play_times = play_times[i].get("play_times")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisGroup, inf_materialPlayCount,sql_play_times))
        # self.assertEqual(inf_materialPlayCount,sql_play_times,"test_playListplayTimes,数据对不上" )
        statisDates = 'statisGroup'
        com_materialPlayCount = 'materialPlayCount'
        com_play_times = 'play_times'
        comparison_1(self, inf_materialPlayCount, sql_play_times, statisDates, com_materialPlayCount, com_play_times)

    # 播放趋势图_请求屏数
    @skip_dependon(depend="test_analysisPlayLen")
    def test_playListrequestScreenCount(self):
        print("开始测试test_playListrequestScreenCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_requestScreenCount = self.res_date['data']

        # 执行SQL查询
        sql = "SELECT transaction_date,SUM(request_screens) AS request_screens FROM	ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}'AND '{2}' GROUP BY transaction_date;".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_request_screens = self.pymysqlcursor.fetchall()

        print("请求屏数对比")
        # for i in range(len(requestScreenCount)):
        #     statisGroup = requestScreenCount[i].get("statisGroup")
        #     inf_requestScreenCount = requestScreenCount[i].get("requestScreenCount")
        #     sql_request_screens = request_screens[i].get("request_screens")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisGroup, inf_requestScreenCount,sql_request_screens))
        # self.assertEqual(inf_requestScreenCount,sql_request_screens,"test_playListrequestScreenCount,数据对不上" )
        statisDates = 'statisGroup'
        com_requestScreenCount = 'requestScreenCount'
        com_request_screens = 'request_screens'
        comparison_1(self, inf_requestScreenCount, sql_request_screens, statisDates, com_requestScreenCount, com_request_screens)

    # 播放趋势图_异常屏数
    @skip_dependon(depend="test_analysisPlayLen")
    def test_playListerrorScreenCount(self):
        print("开始测试test_playListerrorScreenCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_errorScreenCount = self.res_date['data']

        # 执行SQL查询
        sql = "SELECT ssu.create_time, (SUM(DISTINCT ssu.screen_count) - SUM(DISTINCT spu.play_screens)) AS errorScreenCounts FROM screen_summary AS ssu LEFT JOIN ssp_play_summary AS spu ON ssu.create_time = spu.transaction_date WHERE	ssu.media_provider_id = {0} AND (ssu.create_time BETWEEN '{1}' AND  '{2}') GROUP BY create_time;".format( media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_errorScreenCounts = self.pymysqlcursor.fetchall()

        print("异常屏对比")
        # for i in range(len(errorScreenCount)):
        #     statisGroup = errorScreenCount[i].get("statisGroup")
        #     inf_errorScreenCount = errorScreenCount[i].get("errorScreenCount")
        #     sql_errorScreenCount = sql_errorScreenCounts[i].get("errorScreenCounts")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisGroup, inf_errorScreenCount,sql_errorScreenCount))
        # self.assertEqual(inf_errorScreenCount,sql_errorScreenCount,"test_playListerrorScreenCount,数据对不上" )
        statisDates = 'statisGroup'
        com_errorScreenCount = 'errorScreenCount'
        com_errorScreenCounts = 'errorScreenCounts'
        comparison_1(self, inf_errorScreenCount, sql_errorScreenCounts, statisDates, com_errorScreenCount, com_errorScreenCounts)

    # 播放趋势图_播放异常次数
    @skip_dependon(depend="test_analysisPlayLen")
    def test_playListmaterialFailCount(self):
        print("开始测试test_playListmaterialFailCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialFailCount = self.res_date['data']

        # 执行SQL查询
        sql = "SELECT transaction_date,SUM(play_fail_times) AS play_fail_times FROM	ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}'AND '{2}' GROUP BY transaction_date;".format( media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_play_fail_times = self.pymysqlcursor.fetchall()

        print("播放异常对比")
        # for i in range(len(materialFailCount)):
        #     statisGroup = materialFailCount[i].get("statisGroup")
        #     inf_materialFailCount = materialFailCount[i].get("materialFailCount")
        #     sql_play_fail_times = play_fail_times[i].get("play_fail_times")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisGroup, inf_materialFailCount,sql_play_fail_times))
        # self.assertEqual(inf_materialFailCount,sql_play_fail_times,"test_playListmaterialFailCount,数据对不上" )
        statisDates = 'statisGroup'
        com_materialFailCount = 'materialFailCount'
        com_play_fail_times = 'play_fail_times'
        comparison_1(self, inf_materialFailCount, sql_play_fail_times, statisDates, com_materialFailCount, com_play_fail_times)

    # 播放趋势图_请求广告次数
    @skip_dependon(depend="test_analysisPlayLen")
    def test_playListrequestCount(self):
        print("开始测试test_playListrequestCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + sspFinanceAnalysisPlayList,data = FinanceAnalysisPlayList_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_requestCount = self.res_date['data']

        # 执行SQL查询
        sql = "SELECT transaction_date,SUM(request_times) AS request_times FROM	ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}'AND '{2}' GROUP BY transaction_date;".format( media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_request_times = self.pymysqlcursor.fetchall()

        print("请求广告次数对比")
        # for i in range(len(requestCount)):
        #     statisGroup = requestCount[i].get("statisGroup")
        #     inf_requestCount = requestCount[i].get("requestCount")
        #     sql_request_times = request_times[i].get("request_times")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisGroup, inf_requestCount,sql_request_times))
        # self.assertEqual(inf_requestCount,sql_request_times,"test_playListrequestCount,数据对不上" )
        statisDates = 'statisGroup'
        com_requestCount = 'requestCount'
        com_request_times = 'request_times'
        comparison_1(self, inf_requestCount, sql_request_times, statisDates, com_requestCount, com_request_times)


if __name__ == "__main__":
    unittest.main()