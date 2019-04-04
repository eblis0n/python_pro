# encoding=utf-8
import unittest
from utils.methods import set_dp_interface as set_dp_interface
import pymysql
import os
import configparser as cparser
import openpyxl


############################### 配置文件的地址######################
base_dr = str(os.path.dirname(os.path.dirname(__file__)))
bae_idr = base_dr.replace('\\', '/')
# 配置文件的地址
file_path = bae_idr + "/config.ini"
conf = cparser.ConfigParser()
conf.read(file_path)

#######################读取配置文件上###################################
mysqlHost = conf.get("mysqlconf", "host")   # 读取配置文件上数据库主机号
mysqlPort = conf.get("mysqlconf", "port")    # 读取配置文件数据库的端口号
mysqlUser = conf.get("mysqlconf", "user")    # 读取配置文件数据库的账号
mysqlPassword = conf.get("mysqlconf", "password")   # 读取配置文件数据库的密码
mysqlName = conf.get("mysqlconf", "db_name")   # 读取配置文件数据库的名字
wbook = openpyxl.load_workbook(conf.get("dmpexcel", "excel_path")) # 读取配置文件的Excel文档地址
table2 = wbook[conf.get("dmpexcel", "table")] # 读取配置文件的Excel文档所使用的用例table

#######################配置Excel的用例地址#################################
#
# wbook=openpyxl.load_workbook(bae_idr + "/testexcel/dmptest3.xlsx")
# table2=wbook['Test_Case2']
# # print(table2.title)

#######################根据Excel读取本接口对应的key#################################
begtest=[]
endtest=[]

for i in range(1,int(table2.max_column)):
    if table2["A"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

for j in range(1, len(begtest)):
    for k in range(2, int(begtest[j] + 1)):
        # print("excel:" + table2["C" + str(k)].value)
        if os.path.basename(__file__)[0:-3] == table2["C" + str(k)].value:

            # 接口参数
            planentry_url = table2["D" + str(k)].value
            totalsummary_address = table2["E" + str(k)].value
            totalsummary_data = table2["F" + str(k)].value

            # 数据库条件
            create_time = table2["I" + str(k)].value
            startDate = table2["J" + str(k)].value
            endDate = table2["K" + str(k)].value
            break
########################################################


class test_analysis_summary (unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pymysqlconn = pymysql.connect(host=mysqlHost, port=int(mysqlPort), user=mysqlUser, passwd=mysqlPassword, db=mysqlName,
                           charset="utf8",cursorclass = pymysql.cursors.DictCursor)
        cls.pymysqlcursor = cls.pymysqlconn.cursor()

        # 接口查询
        cls.res_date = set_dp_interface(url=planentry_url + totalsummary_address,data = totalsummary_data)

        print("dmp_analysis_summary接口测试开始")

    @classmethod
    def tearDownClass(cls):
        print("测试结束")
        cls.pymysqlconn.close()

    # top 10 广告主数
    def test_playTimeLen(self):
        print("开始测试test_playTimeLen")
        # 执行接口查询
        # self.res_date = set_dp_interface(url=interface.get("queryentry_url") + queryentrySspSummary,data = sspSummary_data)
        print("接口返回结果:" + str(self.res_date))
        inf_playTimeLen = self.res_date['data']
        # 执行SQL查询
        sql = "SELECT COUNT(DISTINCT dsp_id) AS play_time_len  FROM dsp_play_summary WHERE transaction_date BETWEEN '{0}' and '{1}';".format(
            startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        play_time_len = self.pymysqlcursor.fetchall()
        sql_play_time_len = play_time_len[0].get("play_time_len")
        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(len(inf_playTimeLen), len(sql_play_time_len)))
        self.assertEquals(str(len(inf_playTimeLen)), str(len(sql_play_time_len)), "test_playTimeLen数据对不上")


if __name__ == "__main__":
    unittest.main()
