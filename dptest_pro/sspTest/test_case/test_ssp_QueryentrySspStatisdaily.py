# encoding=utf-8

import unittest
from utils.methods import skip_dependon as skip_dependon
from utils.methods import comparison_1 as comparison_1
from utils.methods import set_dp_interface as set_dp_interface
import constants.deploy as deploy
import pymysql
import os
import configparser  as cparser
import openpyxl

############################### 配置文件的地址######################

# base_dr = str(os.path.dirname(os.path.dirname(__file__)))
# bae_idr = base_dr.replace('\\', '/')
#
# file_path = bae_idr + "/config.ini"
#
# conf = cparser.ConfigParser()
#
# conf.read(file_path)
#
#
# # ######################读取配置文件上###################################
# mysqlHost = conf.get("mysqlconf", "host")   # 读取配置文件上数据库主机号
# mysqlPort = conf.get("mysqlconf", "port")    # 读取配置文件数据库的端口号
# mysqlUser = conf.get("mysqlconf", "user")    # 读取配置文件数据库的账号
# mysqlPassword = conf.get("mysqlconf", "password")   # 读取配置文件数据库的密码
# mysqlName = conf.get("mysqlconf", "db_name")   # 读取配置文件数据库的名字
# wbook = openpyxl.load_workbook(conf.get("excel", "excel_path")) # 读取配置文件的Excel文档地址
# table2 = wbook[conf.get("excel", "table")] # 读取配置文件的Excel文档所使用的用例table
#######################配置Excel的用例地址#################################
# 直接读取不利于后期修改Excel文档存放变更
# wbook=openpyxl.load_workbook(bae_idr + "/testexcel/ssptest.xlsx")
# table2=wbook['Test_Case2']

#######################从deploy里面读取配置文件内容#################################

mysqlHost = deploy.mysqlHost
mysqlPort = deploy.mysqlPort
mysqlUser = deploy.mysqlUser
mysqlPassword = deploy.mysqlPassword
mysqlName = deploy.mysqlName
wbook = deploy.wbook
table2 = deploy.table2

#######################根据Excel读取本接口对应的key#################################
begtest=[]
endtest=[]

for i in range(1,int(table2.max_column)):
    if table2["C"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

for j in range(1, len(begtest)):
    for k in range(2, int(begtest[j] + 1)):
        # print("excel:" + table2["C" + str(k)].value)
        if os.path.basename(__file__)[0:-3] == table2["D" + str(k)].value:

            # 接口参数
            queryentry_url = table2["E" + str(k)].value # 接口路径
            queryentrySspStatisdaily = table2["F" + str(k)].value # 接口地址
            sspStatisdaily_data = table2["G" + str(k)].value # 请求参数

            # 数据库条件
            media_provider_id = table2["H" + str(k)].value
            # create_time = table2["I" + str(k)].value
            # transaction_date = table2["H" + str(k)].value
            startDate = table2["I" + str(k)].value
            endDate = table2["J" + str(k)].value
            limits = table2["k" + str(k)].value
            break
########################################################

# 对应的接口：/queryentry/ssp/statisdaily
class test_queryentry_ssp_statisdaily (unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pymysqlconn = pymysql.connect(host=mysqlHost, port=int(mysqlPort), user=mysqlUser, passwd=mysqlPassword, db=mysqlName,
                           charset="utf8",cursorclass = pymysql.cursors.DictCursor)
        cls.pymysqlcursor = cls.pymysqlconn.cursor()

        # 接口
        # cls.res_date = set_dp_interface(url=queryentry_url + queryentrySspStatisdaily,data = sspStatisdaily_data)

        print("queryentry_ssp_statisdaily接口测试开始\n")

    @classmethod
    def tearDownClass(cls):
        print("测试结束")
        cls.pymysqlconn.close()

    # 素材播放趋势图_播放次数统计
    def test_dataLen(self):
        print("开始测试test_dataLen")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspStatisdaily,data = sspStatisdaily_data)
        # print("接口返回结果:" + str(self.res_date))
        dataLen =  self.res_date['data']
        # print(len(dataLen))
        # 执行SQL查询
        # self.sql_screenCount = get_media_screenCount(2028, '2019-02-14',conn = self.pymysqlconn )
        sql = "SELECT DISTINCT (transaction_date) FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}';".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_dataLen=self.pymysqlcursor.fetchall()
        # print(len(sql_dataLen))
        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(len(dataLen),len(sql_dataLen)))
        self.assertEqual(len(dataLen), len(sql_dataLen),"test_dataLen,数据对不上" )

    # 素材播放趋势图_播放次数统计_终端数量
    @skip_dependon(depend="test_dataLen")
    def test_validScreenCount(self):
        print("开始测试test_validScreenCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspStatisdaily,data = sspStatisdaily_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_validScreenCount =  self.res_date['data']

        # 执行SQL查询
        sql = "SELECT DISTINCT transaction_date,SUM(play_screens) AS play_screens FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}' GROUP BY transaction_date;".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_play_screens=self.pymysqlcursor.fetchall()
        print("屏数量比对")
        # for i in range(len(inf_validScreenCount)):
        #     statisDate = inf_validScreenCount[i].get("statisDate")
        #     com_validScreenCount = inf_validScreenCount[i].get("validScreenCount")
        #     com_play_screens = sql_play_screens[i].get("play_screens")
        #
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisDate, com_validScreenCount,com_play_screens))
        # self.assertEqual(str(com_validScreenCount),str(com_play_screens),"test_validScreenCount,数据对不上" )

        # 调用methods中的方法，进行判断
        statisDates = 'statisDate'
        com_validScreenCount = 'validScreenCount'
        com_play_screens = 'play_screens'
        comparison_1(self, inf_validScreenCount, sql_play_screens, statisDates, com_validScreenCount, com_play_screens)

    # 素材播放趋势图_播放次数统计_播放次数
    @skip_dependon(depend="test_dataLen")
    def test_playTimes(self):
        print("开始测试test_playTimes")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspStatisdaily,data = sspStatisdaily_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_playTimes =  self.res_date['data']

        # 执行SQL查询
        sql = "SELECT DISTINCT transaction_date,SUM(play_times) AS play_times FROM	ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}' GROUP BY transaction_date;".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_play_times=self.pymysqlcursor.fetchall()
        print("播放次数比对")
        # for i in range(len(inf_playTimes)):
        #     statisDate = inf_playTimes[i].get("statisDate")
        #     com_playTimes = inf_playTimes[i].get("playTimes")
        #     com_play_times = sql_play_times[i].get("play_times")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisDate, com_playTimes,com_play_times))
        # self.assertEqual(str(com_playTimes),str(com_play_times),"test_playTimes,数据对不上" )
        statisDates = 'statisDate'
        com_playTimes = 'playTimes'
        com_play_times = 'play_times'
        comparison_1(self, inf_playTimes, sql_play_times, statisDates, com_playTimes, com_play_times)

    # 素材播放趋势图_播放时长
    @skip_dependon(depend="test_dataLen")
    def test_playDuration(self):
        print("开始测试test_playDuration")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspStatisdaily,data = sspStatisdaily_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_playDuration =  self.res_date['data']
        # print(len(inf_playDuration))

        # 执行SQL查询
        sql = "SELECT  DISTINCT transaction_date,SUM(play_duration) AS play_duration FROM	ssp_play_summary WHERE  ssp_id = {0} AND (transaction_date BETWEEN '{1}' AND '{2}') GROUP BY transaction_date;".format(media_provider_id, startDate, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        sql_play_duration=self.pymysqlcursor.fetchall()
        # print(len(sql_playDuration))
        print("播放时长对比")
        # for i in range(len(inf_playDuration)):
        #     statisDate = inf_playDuration[i].get("statisDate")
        #     com_playDurations = inf_playDuration[i].get("playDuration")
        #     com_play_durations = sql_play_duration[i].get("play_duration")
        #     print('日期: {0}, 接口返回结果：{1} , 数据库查询结果：{2}' .format(statisDate, com_playDurations,com_play_durations))
        # self.assertEqual(str(com_playDurations),str(com_play_durations),"test_playDuration,数据对不上" )
        statisDates = 'statisDate'
        com_playDurations = 'playDuration'
        com_play_durations = 'play_duration'
        comparison_1(self,inf_playDuration,sql_play_duration,statisDates,com_playDurations,com_play_durations)







if __name__ == "__main__":
    unittest.main()