# encoding=utf-8

import unittest
from utils.methods import set_dp_interface as set_dp_interface
import constants.deploy as deploy
import openpyxl
import pymysql
import os
import configparser  as cparser

############################### 配置文件的地址######################

# base_dr = str(os.path.dirname(os.path.dirname(__file__)))
# bae_idr = base_dr.replace('\\', '/')
#
# file_path = bae_idr + "/config.ini"
#
# conf = cparser.ConfigParser()
#
# conf.read(file_path)
#
#
# # ######################读取配置文件上###################################
# mysqlHost = conf.get("mysqlconf", "host")   # 读取配置文件上数据库主机号
# mysqlPort = conf.get("mysqlconf", "port")    # 读取配置文件数据库的端口号
# mysqlUser = conf.get("mysqlconf", "user")    # 读取配置文件数据库的账号
# mysqlPassword = conf.get("mysqlconf", "password")   # 读取配置文件数据库的密码
# mysqlName = conf.get("mysqlconf", "db_name")   # 读取配置文件数据库的名字
# wbook = openpyxl.load_workbook(conf.get("excel", "excel_path")) # 读取配置文件的Excel文档地址
# table2 = wbook[conf.get("excel", "table")] # 读取配置文件的Excel文档所使用的用例table
#######################配置Excel的用例地址#################################
# 直接读取不利于后期修改Excel文档存放变更
# wbook=openpyxl.load_workbook(bae_idr + "/testexcel/ssptest.xlsx")
# table2=wbook['Test_Case2']

#######################从deploy里面读取配置文件内容#################################

mysqlHost = deploy.mysqlHost
mysqlPort = deploy.mysqlPort
mysqlUser = deploy.mysqlUser
mysqlPassword = deploy.mysqlPassword
mysqlName = deploy.mysqlName
wbook = deploy.wbook
table2 = deploy.table2

#######################根据Excel读取本接口对应的key#################################
begtest=[]
endtest=[]

for i in range(1,int(table2.max_column)):
    if table2["C"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

for j in range(1, len(begtest)):
    for k in range(2, int(begtest[j] + 1)):
        # print("excel:" + table2["C" + str(k)].value)
        if os.path.basename(__file__)[0:-3] == table2["D" + str(k)].value:

            # 接口参数
            queryentry_url = table2["E" + str(k)].value # 接口路径
            queryentrySspSummary = table2["F" + str(k)].value # 接口地址
            sspSummary_data = table2["G" + str(k)].value # 请求参数

            # 数据库条件
            media_provider_id = table2["H" + str(k)].value
            # create_time = table2["I" + str(k)].value
            # transaction_date = table2["H" + str(k)].value
            startDate = table2["I" + str(k)].value
            endDate = table2["J" + str(k)].value
            limits = table2["k" + str(k)].value
            break
########################################################


# 对应财务模块接口：/finance/request/summary
class test_queryentry_ssp_summary (unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.pymysqlconn = pymysql.connect(host=mysqlHost, port=int(mysqlPort), user=mysqlUser, passwd=mysqlPassword, db=mysqlName,
                           charset="utf8",cursorclass = pymysql.cursors.DictCursor)
        cls.pymysqlcursor = cls.pymysqlconn.cursor()
        # 接口
        # cls.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary,data = sspSummary_data)

        print("queryentry_ssp_summary接口测试开始\n")

    @classmethod
    def tearDownClass(cls):
        print("测试结束")
        cls.pymysqlconn.close()


    # 终端总数
    def test_screenCount(self):
        print("开始测试test_screenCount")
        # 执行接口查询
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_screenCount =  self.res_date['data'].get("screenCount")
        # 执行SQL查询
        sql = "SELECT SUM(screen_count) as screen_count FROM screen_summary WHERE media_provider_id = {0} AND create_time BETWEEN '{1}' AND '{2}'  GROUP BY create_time  ORDER BY create_time DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        screen_count=self.pymysqlcursor.fetchall()
        sql_screen_count = screen_count[0].get("screen_count")
        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}' .format(str(inf_screenCount),str(sql_screen_count)))
        self.assertEquals(str(inf_screenCount),str(sql_screen_count),"test_screenCount数据对不上" )

    # 有效总数（有播放广告的设备数）
    def test_validScreenCount(self):
        print("开始测试test_validScreenCount")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_validScreenCount =  self.res_date['data'].get("validScreenCount")

        # 执行SQL查询
        sql = "SELECT SUM(play_screens)as play_screens FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        request_screens=self.pymysqlcursor.fetchall()
        sql_play_screens =  request_screens[0].get("play_screens")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_validScreenCount), str(sql_play_screens)))
        self.assertEquals(str(inf_validScreenCount), str(sql_play_screens), "test_validScreenCount数据对不上")

    # 播放素材数量（有被播放的素材）
    def test_materialCount(self):
        print("开始测试test_materialCount")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialCount =  self.res_date['data'].get("materialCount")

        # 执行SQL查询
        sql = "SELECT SUM(play_material_count) as play_material_count FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        material_count=self.pymysqlcursor.fetchall()
        slq_play_material_count =  material_count[0].get("play_material_count")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_materialCount), str(slq_play_material_count)))
        self.assertEquals(str(inf_materialCount), str(slq_play_material_count), "test_materialCount数据对不上")

    # 有效播放次数
    def test_materialPlayCount(self):
        print("开始测试test_materialPlayCount")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialPlayCount = self.res_date['data'].get("materialPlayCount")

        # 执行SQL查询
        sql = "SELECT SUM(play_times) as play_times FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}' GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        play_times = self.pymysqlcursor.fetchall()
        sql_play_times = play_times[0].get("play_times")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_materialPlayCount), str(sql_play_times)))
        self.assertEquals(str(inf_materialPlayCount), str(sql_play_times), "test_materialPlayCount数据对不上")

    # 异常播放（播放失败）
    def test_materialFailCount(self):
        print("开始测试test_materialFailCount")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialFailCount = self.res_date['data'].get("materialFailCount")

        # 执行SQL查询
        sql = "SELECT SUM(play_fail_times) as play_fail_times FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        play_times = self.pymysqlcursor.fetchall()
        sql_play_fail_times = play_times[0].get("play_fail_times")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_materialFailCount), str(sql_play_fail_times)))
        self.assertEquals(str(inf_materialFailCount), str(sql_play_fail_times), "test_materialFailCount数据对不上")

    # 播放时长
    def test_materialPlayDuration(self):
        print("开始测试test_materialPlayDuration")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_materialPlayDuration = self.res_date['data'].get("materialPlayDuration")

        # 执行SQL查询
        sql = "SELECT SUM(play_duration) as play_duration FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        play_duration = self.pymysqlcursor.fetchall()
        sql_play_duration = play_duration[0].get("play_duration")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_materialPlayDuration), str(sql_play_duration)))
        self.assertEquals(str(inf_materialPlayDuration), str(sql_play_duration), "test_materialPlayDuration数据对不上")

    # 营业收入
    @unittest.skip("营业收入前端暂时隐藏")
    def test_income(self):
        print("开始测试test_income")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_income = self.res_date['data'].get("income")

        # 执行SQL查询
        sql = "SELECT SUM(amount) as amount FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '{1}' AND '{2}'  GROUP BY transaction_date  ORDER BY transaction_date DESC LIMIT {3};".format(media_provider_id, startDate, endDate, limits)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        amount = self.pymysqlcursor.fetchall()
        sql_amount = amount[0].get("amount")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_income), str(sql_amount)))
        self.assertEquals(str(inf_income), str(sql_amount), "test_income数据对不上")

    # 累计收益
    @unittest.skip("累计收益前端暂时隐藏")
    def test_allIncome(self):
        print("开始测试test_allIncome")
        self.res_date = set_dp_interface(url=queryentry_url + queryentrySspSummary, data=sspSummary_data)
        # print("接口返回结果:" + str(self.res_date))
        inf_allIncome = self.res_date['data'].get("allIncome")

        # 执行SQL查询
        sql = "SELECT SUM(amount) as allamount FROM ssp_play_summary WHERE ssp_id = {0} AND transaction_date BETWEEN '2019-01-01' and '{1}';".format(media_provider_id, endDate)
        print("数据库查询使用语句：" + sql)
        self.pymysqlcursor.execute(sql)
        allamount = self.pymysqlcursor.fetchall()
        sql_allamount = allamount[0].get("allamount")

        # 判断 接口数据与数据库查询结果
        print('接口返回结果：{0} , 数据库查询结果：{1}'.format(str(inf_allIncome), str(sql_allamount)))
        self.assertEquals(str(inf_allIncome), str(sql_allamount), "test_allIncome数据对不上")


if __name__ == "__main__":
    unittest.main()