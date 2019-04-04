# encoding=utf-8

import os
import configparser as cparser
import openpyxl

############################### 配置文件的地址######################

base_dr = str(os.path.dirname(os.path.dirname(__file__)))
bae_idr = base_dr.replace('\\', '/')

file_path = bae_idr + "/config.ini"

conf = cparser.ConfigParser()

conf.read(file_path)


#######################测试读取配置文件上###################################
# mysqlHost = conf.get("mysqlconf", "host")   # 读取配置文件上数据库主机号
# mysqlPort = conf.get("mysqlconf", "port")    # 读取配置文件数据库的端口号
# mysqlUser = conf.get("mysqlconf", "user")    # 读取配置文件数据库的账号
# mysqlPassword = conf.get("mysqlconf", "password")   # 读取配置文件数据库的密码
# mysqlName = conf.get("mysqlconf", "db_name")   # 读取配置文件数据库trans_db
# mysqlName1 = conf.get("mysqlconf", "db_name1")   # 读取配置文件数据库ad_account
# wbook = openpyxl.load_workbook(conf.get("excel", "ssp_excel_path")) # 读取配置文件的Excel文档地址
# dmpWbook = openpyxl.load_workbook(conf.get("excel", "dmp_excel_path")) # 读取配置文件的Excel文档地址
# table2 = wbook[conf.get("excel", "table")] # 读取配置文件的Excel文档所使用的用例table

#######################online读取配置文件上###################################
mysqlHost = conf.get("onlinemysqlconf", "host")   # 读取配置文件上数据库主机号
mysqlPort = conf.get("onlinemysqlconf", "port")    # 读取配置文件数据库的端口号
mysqlUser = conf.get("onlinemysqlconf", "user")    # 读取配置文件数据库的账号
mysqlPassword = conf.get("onlinemysqlconf", "password")   # 读取配置文件数据库的密码
mysqlName = conf.get("onlinemysqlconf", "db_name")   # 读取配置文件数据库trans_db
mysqlName1 = conf.get("onlinemysqlconf", "db_name1")   # 读取配置文件数据库ad_account
wbook = openpyxl.load_workbook(conf.get("onlineexcel", "ssp_excel_path")) # 读取配置文件的Excel文档地址
dmpWbook = openpyxl.load_workbook(conf.get("onlineexcel", "dmp_excel_path")) # 读取配置文件的Excel文档地址
table2 = wbook[conf.get("onlineexcel", "table")] # 读取配置文件的Excel文档所使用的用例table






