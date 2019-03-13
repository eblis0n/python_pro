# -*- coding: utf-8 -*-

import os,sys
import openpyxl
import unittest
import shutil
# def file_name(file_dir):
#     for root, dirs, files in os.walk(file_dir):
#         print(root)  # 当前目录路径
#         print(dirs)  # 当前路径下所有子目录
#         print(files)  # 当前路径下所有非目录子文件


base_dr = str(os.path.dirname(os.path.dirname(__file__)))
bae_idr = base_dr.replace('\\', '/')
cur_path = os.path.dirname(os.path.realpath(__file__))

caseName="testcase"
aa=os.path.join(bae_idr, caseName)

print(aa)
wbook=openpyxl.load_workbook(bae_idr + "/testexcel/dmptest3.xlsx")
table2=wbook['Test_Case2']
#

begtest=[]
endtest=[]


for i in range(1,int(table2.max_column)):
    if table2["A"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

for j in range(1, len(begtest)):
    for k in range(2, int(begtest[j] + 1)):
        # print("excel:" + table2["C" + str(k)].value)
        print(os.path.basename(sys.argv[0]))
        if os.path.basename(sys.argv[0])[0:-3] == table2["C" + str(k)].value:
            print(str(k))
            vv = str(k)
            totalsummary_address = table2["E" + str(k)].value
            print(totalsummary_address)



# def aaa():
#     begtest=[]
#     endtest=[]
#     vv = []
#
#     for i in range(1,int(table2.max_column)):
#         if table2["A"+str(i)].value=="go":
#             begtest.append(i)
#
#         if table2["L"+str(i)].value=="end":
#             endtest.append(i)
#
#     for j in range(1, len(begtest)):
#         for k in range(2, int(begtest[j] + 1)):
#             # print("excel:" + table2["C" + str(k)].value)
#             if os.path.basename(sys.argv[0])[0:-3] == table2["C" + str(k)].value:
#                 print(str(k))
#                 vv = str(k)
#                 # totalsummary_address = table2["E" + str(k)].value
#     return vv