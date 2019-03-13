# -*- coding: utf-8 -*-

import os,sys
import openpyxl
import unittest
import shutil
# def file_name(file_dir):
#     for root, dirs, files in os.walk(file_dir):
#         print(root)  # 当前目录路径
#         print(dirs)  # 当前路径下所有子目录
#         print(files)  # 当前路径下所有非目录子文件


base_dr = str(os.path.dirname(os.path.dirname(__file__)))
bae_idr = base_dr.replace('\\', '/')
cur_path = os.path.dirname(os.path.realpath(__file__))

caseName="testcase"
aa=os.path.join(bae_idr, caseName)

print(aa)
wbook=openpyxl.load_workbook(bae_idr + "/testexcel/dmptest3.xlsx")
table2=wbook['Test_Case2']
#

# def file_name(file_dir):
#     L = []
#     for root, dirs, files in os.walk(file_dir):
#         for file in files:
#             if os.path.join(file)[0:4] == "test":
#                 if os.path.splitext(file)[1] == '.py':
#                     L.append(os.path.join( file)[0:-3])
#     return L


    # 其中os.path.splitext()函数将路径拆分为文件名+扩展名

# def table_row(filename):
#     begtest = []
#     endtest = []
#     for i in range(1, int(table2.max_column)):
#        if table2["A" + str(i)].value == "go":
#            begtest.append(i)
#
#        if table2["L" + str(i)].value == "end":
#            endtest.append(i)

    # print(len(begtest))
    #
    #
    # ss = []
    # for i in range(0,len(filename)):
    #    # print("casename:" + file_name(aa)[i])
    #    print("第" + str(i) + "条用例")
    #    for j in range(1,len(begtest)):
    #        for k in range(2, int(begtest[i] + 1)):
    #            # print("excel:"+table2["C"+str(k)].value)
    #            if filename[i] == table2["C"+str(k)].value:
    #                print("当前行号:" + str(k))

       # print(table2["A" + str(k)].value)
       # print(file_name(aa)[i])
       # bb = str(file_name(aa)[i] + '.py')
       # print(bb)
       # ss.append(str(k))
    # return ss

# sss = table_row(file_name(aa))
#
# if filename[i] == table2["C"+str(k)].value:
#     vv = str(k)
#     totalsummary_address = table2["E"+str(vv)].value
#     print(totalsummary_address)




# print (os.path.basename(sys.argv[0]))






# vv  = []
# for root, dirs, files in os.walk(aa):
#     # print(files)
#     for file in  files :
#         print(os.path.join(file)[0:4])
#         if os.path.join(file)[0:4] == "test":
#             # vv.append(os.path.join(file))
#             print(os.path.join(file))

# print(vv)


#
# print(file_name(aa))
# # print(len(file_name(aa)))
#
#
#
#
#
# begtest=[]
# endtest=[]
#
# for i in range(1,int(table2.max_column)):
#     if table2["A"+str(i)].value=="go":
#         begtest.append(i)
#
#     if table2["L"+str(i)].value=="end":
#         endtest.append(i)
# if len(begtest)==0:#检查是否有可用的用例
#     print('没有可执行的用例')
#     quit()
# if len(begtest)!=len(endtest):
#     print("用例有误,请检查各单元用例的开始和结束是否一致")
#     quit()
# print("测试用例加载完成")
#
# # print(len(begtest))
#
# ss = []
# for i in range(0,len(file_name(aa))):
#     # print("casename:" + file_name(aa)[i])
#     print("第" + str(i) + "条用例")
#     for j in range(1,len(begtest)):
#         for k in range(2, int(begtest[i] + 1)):
#
#             print("excel:"+table2["C"+str(k)].value)
#
#             if file_name(aa)[i] == table2["C"+str(k)].value:
#                 print("当前行号:" + str(k))
            #     # print(table2["A" + str(k)].value)
            #     # print(file_name(aa)[i])
            #     bb = str(file_name(aa)[i] + '.py')
            #     print(bb)
            #     ss.append(str(file_name(aa)[i] + '.py'))
            # k = k + 1

# for i in range(0,len(ss)):
#
#     if not os.path.exists(aa +"/test_case"):
#         os.makedirs(aa +"/test_case")
#     # for root, dirs, files in os.walk(aa):
#     #     for file in  files:
#     #         if os.path.join(file) == ss[i]:
#     #             shutil.copy(os.path.join(file), 'test_case')
#
#     shutil.copy(aa + "/" + ss[i], aa +"/test_case")



# print(ss)

# for l in range(0,len(ss)):
#
#     discover = unittest.defaultTestLoader.discover(aa,
#                                                    pattern=ss[l],
#                                                    top_level_dir=None)
#     print(discover)

# discover = unittest.defaultTestLoader.discover(aa,pattern=ss,top_level_dir=None)
# print(discover)

# for i in range(1,len(begtest)):
#     for j in range(2,int(begtest[i]+1)):
#         print("zhengque")
#
# for b in range(0,int(len(file_name(aa)))):
#
#     if file_name(aa)[b] == table2["C"+str(j)].value:
#         print(table2["A" + str(j)].value)
#         print(file_name(aa)[b])
    #     bb = str(file_name(aa)[b] + '.py')
    #     print(bb)
    #     # ss = []
    #     # ss.append(str(file_name(aa)[b] + '.py'))
    #     # print(bb)
    #     # discover = unittest.defaultTestLoader.discover(aa,
    #     #                                                pattern=bb,
    #     #                                                top_level_dir=None)
    #     # print(discover)
    #
    # else:
    #     print("错误")

# begtest=[]
# endtest=[]
#
# for i in range(1,int(table2.max_column)):
#     if table2["A"+str(i)].value=="go":
#         begtest.append(i)
#
#     if table2["L"+str(i)].value=="end":
#         endtest.append(i)
#
# for j in range(1, len(begtest)):
#     for k in range(2, int(begtest[i] + 1)):
#         # print("excel:" + table2["C" + str(k)].value)
#         if os.path.basename(sys.argv[0])[0:-3] == table2["C" + str(k)].value:
#             totalsummary_address = table2["E"+str(k)].value

# for root, dirs, files in os.walk(aa):
#    for file in files:
#        if os.path.join(file)[0:4] == "test":
#            if os.path.join(file)[-3:] == '.py':
#                if os.path.join(file) == os.path.basename(sys.argv[0]):
#
#                    print(os.path.basename(sys.argv[0])[0:-3])
#
#
#                    break




begtest=[]
endtest=[]

for i in range(1,int(table2.max_column)):
    if table2["A"+str(i)].value=="go":
        begtest.append(i)

    if table2["L"+str(i)].value=="end":
        endtest.append(i)

# print(len(begtest))


for j in range(1,len(begtest)):
    for k in range(2, int(begtest[j] + 1)):

        # print("excel:"+table2["C"+str(k)].value)
        if os.path.basename(sys.argv[0])[0:-3] == table2["C"+str(k)].value:

            print(table2["A" + str(k)].value)
            # print(file_name(aa)[i])
            # bb = str(file_name(case_path)[i] + '.py')
            # ss.append(str(file_name(case_path)[i] + '.py'))
        k = k + 1




# def aaa():
#     begtest=[]
#     endtest=[]
#     vv = []
#
#     for i in range(1,int(table2.max_column)):
#         if table2["A"+str(i)].value=="go":
#             begtest.append(i)
#
#         if table2["L"+str(i)].value=="end":
#             endtest.append(i)
#
#     for j in range(1, len(begtest)):
#         for k in range(2, int(begtest[j] + 1)):
#             # print("excel:" + table2["C" + str(k)].value)
#             if os.path.basename(sys.argv[0])[0:-3] == table2["C" + str(k)].value:
#                 print(str(k))
#                 vv = str(k)
#                 # totalsummary_address = table2["E" + str(k)].value
#     return vv